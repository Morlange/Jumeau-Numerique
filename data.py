import openpyxl as op
from opcua import Client
from product import Gamme_de_produit

LogoAM = "logo-AM.ico"


file_Machine=open('liste_machine.txt', 'r')
l=file_Machine.readlines()
liste_machines = []
liste_zone=[]
for k in l:
    
    a=k.split(";")
    #print(a)
    liste_machines.append(a[0])
    liste_zone.append(a[1][:-1])


#données OF



"""document = xlrd.open_workbook('OF.xls')
nb_feuilles = document.nsheets
OF=[]

#Données des machines

#On récupère les informations concernant la production qui sont dans un excel
for i in range (0,nb_feuilles) :
    feuille = document.sheet_by_index(i)
    nom_produit = feuille.name
    nb_colonnes = feuille.ncols
    nb_lignes = feuille.nrows
    n_OF=[]
    nom_OF=[]
    machine_OF=[]
    debut_OF=[]
    caracteristique=[feuille.cell_value(rowx=0,colx=1),feuille.cell_value(rowx=1,colx=1),feuille.cell_value(rowx=2,colx=1) ]
    for i in range (4,nb_lignes):
        n_OF.append(feuille.cell_value(rowx=i, colx=0))
        nom_OF.append(feuille.cell_value(rowx=i, colx=1))
        machine_OF.append(feuille.cell_value(rowx=i, colx=2))
        debut_OF.append(feuille.cell_value(rowx=i, colx=3))
    OF_produit=[nom_produit, n_OF, nom_OF, machine_OF, debut_OF,caracteristique]
    OF.append(OF_produit)
    OF_produit=[]"""

document = op.load_workbook('OF.xlsx')
nom_produits=document.sheetnames
nombre_produits=len(nom_produits)

#données de la fiche produit : donne une liste OF composée de listes gamme 
OF=[]

#forme gamme : [nom_produit, n_operation, nom_operation, machine,temps_fabrication,caracteristique]
#forme des ordres de fabrication : OF=[[nom_produit, n_operation, nom_operation, machine,temps_fabrication,caracteristique], [nom_produit,   n_operation,  , machine,temps_fabrication,caracteristique], ...]

for i in range (nombre_produits) :
    worksheet = document[nom_produits[i]]
    nombre_ligne = worksheet.max_row
    n_operation=[]
    nom_operation=[]
    machine=[]
    temps_fabrication=[]
    #ajout des etapes de fabrication
    for j in range (5,nombre_ligne+1):
        n_operation.append(worksheet.cell(row=j, column=1).value)
        nom_operation.append(worksheet.cell(row=j,column=2).value)
        machine.append(worksheet.cell(row=j,column=3).value)
        temps_fabrication.append(worksheet.cell(row=j,column=4).value)
        cout = worksheet.cell(1,2).value
        resistance = worksheet.cell(2,2).value
        temps = worksheet.cell(3,2).value
        gamme=[]
    gamme.append(nom_produits[i])
    gamme.append(n_operation)
    gamme.append(nom_operation)
    gamme.append(machine)
    gamme.append(temps_fabrication)
    gamme.append([cout, resistance, temps])
    OF.append(gamme)

#forme des odres de fabrication : OF=[[nom_produit, n_OF, nom_OF, machine_OF, debut_OF,caracteristique], [nom_produit, n_OF, nom_OF, machine_OF, debut_OF,caracteristique], ...]
Gammes = []
for i in range (nombre_produits):
    Gammes.append(Gamme_de_produit(OF[i][0],OF[i][2],OF[i][3],OF[i][4]))
#print(OF)
    

#données pour analyse 
TRS=[50,20,30]
Images_TRS = []
Temps_de_production=[46,6,23]
Nombre_pieces=[12,10,7]
LogoAM = "Logo_AM.png"
Images = []
val = 120
taille = (int(4/3*val),val)

#données pour serveur
Serveur_ON = False
url = "opc.tcp://127.0.0.1:4880"
client = Client(url)

#Données couleurs
color_bg1 = "#333333"
color_bg = "#1e1e1e"
color_fg = "#dcdcaa"