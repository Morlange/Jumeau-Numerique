import tkinter as tk
import openpyxl as op
import tkinter
from tkinter import *
from tkinter import ttk
from data import color_bg1,color_bg,color_fg
from PIL import Image, ImageTk
import Pop_up
# A FAIRE :
# - commenter le code
# - mettre "ajout de ligne"
# - rendre joli
# - machine avec liste dans ajout produit
nombre_operation = 1
Images_Temps = []
info_image = Image.open("info.png").resize((50,50))

def main():
    global nombre_operation
    #ouverture du fichier en lecture
    liste_machine=open('liste_machine.txt', 'r')
    l=liste_machine.readlines()
    machine=[]
    nom_machine=[]
    zone_machine=['Composite', 'Usinage', 'Forge','Metrologie','Assemblage','Bois','Fonderie']
    
    ############# #import du document ####################
    document = op.load_workbook('OF.xlsx')
    nom_produits=document.sheetnames
    nombre_produits=len(nom_produits)

    #données de la fiche produit : donne une liste OF composée de listes gamme 
    OF=[]

    #forme gamme : [nom_produit, n_operation, nom_operation, machine,temps_fabrication,caracteristique]
    #forme des ordres de fabrication : OF=[[nom_produit, n_operation, nom_operation, machine,temps_fabrication,caracteristique], [nom_produit,   n_operation,  , machine,temps_fabrication,caracteristique], ...]

    for i in range (nombre_produits) :
        gamme_produit = []
        nom_produit = nom_produits[i]
        worksheet = document[nom_produits[i]]
        nombre_ligne = worksheet.max_row
        n_operation=[]
        nom_operation=[]
        machine=[]
        temps_fabrication=[]
        #ajout des etapes de fabrication
        for j in range (5,nombre_ligne+1):
            n_operation.append(worksheet.cell(row=j, column=1).value)
            nom_operation.append(worksheet.cell(row=j,column=2).value)
            machine.append(worksheet.cell(row=j,column=3).value)
            temps_fabrication.append(worksheet.cell(row=j,column=4).value)
            cout = worksheet.cell(1,2).value
            resistance = worksheet.cell(2,2).value
            temps = worksheet.cell(3,2).value
            gamme=[]
        gamme.append(nom_produits[i])
        gamme.append(n_operation)
        gamme.append(nom_operation)
        gamme.append(machine)
        gamme.append(temps_fabrication)
        gamme.append([cout, resistance, temps])
        OF.append(gamme)
        


    ############### fenetre #######################

    # creation de l'objet fenetrefenetre
    fen_produit= tk.Tk()

    #Titre de la fenetre
    fen_produit.title('Gamme de fabrication')

    #Taille de la fenetre
    fen_produit.geometry('1300x630')

    #icone de fenetre
    fen_produit.iconbitmap('logo-AM.ico')

    #configuration du fond
    fen_produit.config(background=color_bg1)
    #Ajout d'un titre 
    label_titre = Label(fen_produit, text='Gamme de fabrication', height=1,fg=color_fg,font=('Times new roman', 14),bg=color_bg1) 
    label_titre.grid(row=0,column=1, padx=0, pady=10)

    ################# menu déroulant et tableau ###############
    
    def suppr_produit():
        select = listeCombo.get()
        wb = op.load_workbook('OF.xlsx')
        del wb[select]
        wb.save('OF.xlsx')
        fen_produit.destroy()
        main()

    def action(event):
        # Obtenir l'élément sélectionné
        select = listeCombo.get()

        # Srollbar 
        scrollbar = Scrollbar(fen_produit, orient=VERTICAL)
        scrollbar.grid(row=2, column=3, sticky=N+S)
        
        
        style = ttk.Style(fen_produit)
        style.theme_use("winnative")  # Charger un thème pour pouvoir accéder aux éléments de style
        style.configure("Treeview", background=color_bg1, foreground="white", fieldbackground=color_bg1)
        style.map("Treeview", background=[("selected", "#0066CC")], foreground=[("selected", "white")])
        
        #définir le tableau et joindre avec le mouvement de la scrollbar 
        tableau = ttk.Treeview(fen_produit, columns=('gamme_fabrication', 'nom_operation', 'machine', 'temps_fabrication'),show='headings',yscrollcommand = scrollbar.set)
        scrollbar.config( command = tableau.yview )
        
        #définir chaque colonne 
        tableau.heading ('gamme_fabrication', text='gamme fabrication')
        tableau.heading ('nom_operation', text='nom operation')
        tableau.heading ('machine', text='machine')
        tableau.heading ('temps_fabrication', text='temps fabrication')

        #definition de chaque tableau en fonction des produits
        OFs=[]

        for i in range (nombre_produits) :
            if str(select)==nom_produits[i] :
                #caracteristiques produits 
                caracteristique_produit=OF[i][5]
                for j in range (len(OF[i][1])):
                    OFs.append((OF[i][1][j], OF[i][2][j], OF[i][3][j], OF[i][4][j]))

                #autre 
                for valeur in OFs :
                    tableau.insert('', tkinter.END, values=valeur)
                    
                    
        #afficher 
        tableau['show'] = 'headings'
        tableau.grid(row=2, column=1, padx=10, pady=(0,10))

        #Frame Caractéristiques produits 
        frame_caracteristique = tkinter.LabelFrame(fen_produit, text='Caractéristiques du produit', font =('Times new roman', 12), fg=color_fg, bg=color_bg1, labelanchor='nw')
        frame_caracteristique.grid(row=7,column=1,padx=50, pady=10)
        Label(frame_caracteristique, text='Coût :',fg=color_fg,bg=color_bg1).grid(row=3, column=0, padx=10, pady=20)
        Label(frame_caracteristique, text='Résistance :',fg=color_fg,bg=color_bg1).grid(row=4, column=0, padx=10, pady=20)
        Label(frame_caracteristique, text='Temps de cycle moyen :',fg=color_fg,bg=color_bg1).grid(row=5, column=0, padx=10, pady=20)

        Label(frame_caracteristique, text=caracteristique_produit[0], fg=color_fg, bg=color_bg1).grid(row=3, column=2, padx=10, pady=20)
        Label(frame_caracteristique, text=caracteristique_produit[1], fg=color_fg, bg=color_bg1).grid(row=4, column=2, padx=10, pady=20)
        Label(frame_caracteristique, text=caracteristique_produit[2], fg=color_fg, bg=color_bg1).grid(row=5, column=2, padx=10, pady=20)
        
        
        bouton_suppr_produit = tk.Button (fen_produit, text = 'Supprimer le produit', fg=color_fg, bg=color_bg,command=suppr_produit)
        bouton_suppr_produit.grid(row=8,column=1, padx=40,pady=20)

    #liste_N_OP, liste_nom_OP, liste_machine_OP, liste_temps_fab_OP = ajout_produit()
    # fonction ajouter un produit reliée au bouton "Ajouter"
    
    

    def ouvrir_fen_nouveau_produit():

        def ajout_produit():
        #Dernier bouton qui ajoute véritablement dans l'excel
            
        # Gestion du fichier excel
            global nombre_operation
            wb = op.load_workbook('OF.xlsx')
            wb.create_sheet(title=nom_prod.get())
            sheet = wb[nom_prod.get()]
            sheet['A1'] = 'Coût (€)'
            sheet['B1'] = int(cout.get())
            sheet['A2'] = 'Resistance (MPa)'
            sheet['B2'] = int(resistance.get())
            sheet['A3'] = 'Temps de cycle moyen (min) '
            sheet['B3'] = int(temps.get())
            sheet['A4'] = 'Gamme fabrication'
            sheet['B4'] = 'Nom operation'
            sheet['C4'] = 'Machine'
            sheet['D4'] = 'Temps fabrication'
            
            for i in range(nombre_operation):
                a=liste_N_OP[i].get()
                b=liste_nom_OP[i].get()
                c=liste_machine_OP[i].get()
                d=liste_temps_fab_OP[i].get()

                sheet['A{}'.format(i+5)] = a
                sheet['B{}'.format(i+5)] = b
                sheet['C{}'.format(i+5)] = c
                sheet['D{}'.format(i+5)] = d
            
            wb.save('OF.xlsx')
            nouvelle_operation = 1
                
                
                #return (liste_N_OP, liste_nom_OP, liste_machine_OP, liste_temps_fab_OP)
            
        def ajout_produit_bouton():
            global Images_Temps,info_image
            try : 
                ajout_produit()
            except KeyError:
                
                Images_Temps = [Pop_up.main("Veuillez mettre un nom pour le produit",info = True)]

            fenetre.destroy()
            ouvrir_fen_nouveau_produit()
        
        def fermer_ajout_produit():
            global nouvelle_operation
            fenetre.destroy()
            fen_produit.destroy()
            nouvelle_operation = 1
            main()
        
        

        fenetre = tk.Tk()
        fenetre.title('Ajouter un nouveau produit')
        fenetre.geometry('600x300')
        fenetre.iconbitmap('logo-AM.ico')
        tk.Label(fenetre, text="Nom du produit").grid(row=0)
        tk.Label(fenetre, text="Coût (€)").grid(row=1)
        tk.Label(fenetre, text="Résistance (MPa)").grid(row=2)
        tk.Label(fenetre, text="Temps de cycle (min)").grid(row=3)
        tk.Label(fenetre, text="Gamme de fabrication :").grid(row=4)
        tk.Label(fenetre, text="N° operation").grid(row=5,column=0)
        tk.Label(fenetre, text= "Nom operation").grid(row=5,column=1)
        tk.Label(fenetre, text="machine").grid(row=5,column=2)
        tk.Label(fenetre, text="Temps fabrication").grid(row=5,column=3)

        nom_prod = tk.Entry(fenetre)
        cout = tk.Entry(fenetre)
        resistance = tk.Entry(fenetre)
        temps= tk.Entry(fenetre)

        nom_prod.grid(row=0, column=1)
        cout.grid(row=1, column=1)
        resistance.grid(row=2, column=1)
        temps.grid(row=3, column=1) 

        N_OP = tk.Entry(fenetre)
        nom_OP = tk.Entry(fenetre)
        machine_OP=ttk.Combobox(fenetre, values=zone_machine) #liste déroulante des zones
        temps_fab_OP = tk.Entry(fenetre)

        N_OP.grid(row=6, column=0)
        nom_OP.grid(row=6, column=1)
        machine_OP.grid(row=6, column=2)
        temps_fab_OP.grid(row=6, column=3)

        liste_N_OP=[N_OP]
        liste_nom_OP=[nom_OP]
        liste_machine_OP=[machine_OP]
        liste_temps_fab_OP=[temps_fab_OP]

            
        def nouvelle_operation():
            global nombre_operation
            nombre_operation=nombre_operation+1

            N_OP = tk.Entry(fenetre)
            nom_OP = tk.Entry(fenetre)
            machine_OP=ttk.Combobox(fenetre, values=zone_machine) #liste déroulante des zones
            temps_fab_OP = tk.Entry(fenetre)

            N_OP.grid(row=nombre_operation+5, column=0)
            nom_OP.grid(row=nombre_operation+5, column=1)
            machine_OP.grid(row=nombre_operation+5, column=2)
            temps_fab_OP.grid(row=nombre_operation+5, column=3)

            liste_N_OP.append (N_OP)
            liste_nom_OP.append (nom_OP)
            liste_machine_OP.append (machine_OP)
            liste_temps_fab_OP.append (temps_fab_OP)

        bouton_ajout_OP = tk.Button (fenetre, text = 'Ajouter une operation', fg=color_fg, bg=color_bg,command=nouvelle_operation)
        bouton_ajout_OP.grid(row=99, column=2)

        

        tk.Button(fenetre, text='Ajouter le produit',fg=color_fg, bg=color_bg, command=ajout_produit_bouton).grid(row=100, column=2, sticky=tk.W)
        tk.Button(fenetre, text='Fermer',fg=color_fg, bg=color_bg, command=fermer_ajout_produit).grid(row=100,column=1, sticky=tk.W)

        fenetre.bind()

        tk.mainloop()

    # Ajout d'un produit
    bouton = tk.Button (fen_produit, text = 'Ajouter un produit', fg=color_fg, bg=color_bg,command=ouvrir_fen_nouveau_produit)
    bouton.grid(row=1, column=4, padx=0, pady=10)
        
    # Selectionner produit 
    label_choix = Label(fen_produit, text='Choisissez un produit:',fg=color_fg, bg=color_bg1)
    label_choix.grid(row=1, column=0, padx=0, pady=10)
        
    #Création de la Combobox
    listeCombo = ttk.Combobox(fen_produit, values=nom_produits)
        
    #Choisir l'élément qui s'affiche par défaut
    listeCombo.grid(row=1, column=1,padx=0,pady=10)

    listeCombo.bind("<<ComboboxSelected>>", action)   
    
    
        

    #Boucle d'affichage
    fen_produit.mainloop()

if __name__ == "__main__":
    main()